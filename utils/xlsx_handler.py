import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def read_xlsx(file_path):
    """Reads an XLSX file and returns a list of dictionaries."""
    try:
        # Load the Excel file with specific data types for critical columns
        dtype_mapping = {
            "Order ID": str,
            "Tracking Number": str,
            "Zip/Postal Code": str  # Ensure ZIP codes are not treated as numbers
        }

        # Read the Excel file, applying the dtype mappings
        df = pd.read_excel(file_path, dtype=dtype_mapping)

        # Strip unnecessary columns and ensure only required fields are included
        required_columns = [
            "Recipient Name", "Email", "Phone", "Street Line 1",
            "Street Line 2", "City", "State/Province", "Zip/Postal Code",
            "Country", "SKU", "Quantity", "Item Weight", "Item Weight Unit",
            "Order ID", "Tracking Number"
        ]
        df = df[[col for col in df.columns if col in required_columns]]

        # Fill missing values with empty strings (to avoid issues when converting to dict)
        df.fillna("", inplace=True)

        return df.to_dict("records")

    except Exception as e:
        raise ValueError(f"Failed to read XLSX file: {e}")


import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def write_xlsx(file_path, fieldnames, data, sheet_name="Sheet1"):
    """Writes data to an XLSX file with improved formatting."""
    try:
        # Convert data to a DataFrame
        df = pd.DataFrame(data, columns=fieldnames)
        
        # Ensure `Order ID` and `Tracking number` are treated as text
        if "Order ID" in df.columns:
            df["Order ID"] = df["Order ID"].astype(str)
        if "tracking_number" in df.columns:
            df["tracking_number"] = df["tracking_number"].astype(str)
        
        # Write to Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        # Format the Excel file
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        
        # Apply bold formatting to the header
        header_font = Font(bold=True)
        for col_idx, cell in enumerate(ws[1], start=1):
            cell.font = header_font
            # Adjust column width
            column_width = max(len(str(cell.value)), *[
                len(str(ws.cell(row=row_idx, column=col_idx).value or ""))
                for row_idx in range(2, ws.max_row + 1)
            ])
            ws.column_dimensions[get_column_letter(col_idx)].width = column_width + 2

        # Apply alignment to all cells
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Save the updated workbook
        wb.save(file_path)
    except Exception as e:
        raise ValueError(f"Failed to write XLSX file: {e}")